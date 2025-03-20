import os
import logging
import openai
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from typing import Dict, List, Any, Optional, Tuple
from models.gap_definitions import find_gap_examples_by_control, find_gap_examples_by_control_and_status

# Import direct configuration as a fallback
try:
    from config_direct import setup_openai_config
except ImportError:
    # Define the function here as a fallback in case the import fails
    def setup_openai_config():
        logging.getLogger(__name__).info("Using inline fallback configuration")
        openai.api_type = "azure"
        openai.api_base = "https://it-risk-advisory.cognitiveservices.azure.com"
        openai.api_version = "2024-08-01-preview"
        openai.api_key = "6sYzk9nd49SnrWNfxdMsUqeLUnnhfwTOHCnAYVTllARQ1JQxywz0JQQJ99BAACYeBjFXJ3w3AAAAACOGc2jb"
        os.environ["AZURE_DEPLOYMENT_NAME"] = "gpt-4o-it-risk"
        return True

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

class GapGenerator:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.setup_styles()
        # Check if OpenAI is configured properly
        self.check_openai_config()
        self.cache = {}  # Simple cache for API responses
        
    def check_openai_config(self):
        """Verify OpenAI configuration is properly set up"""
        self.logger.debug(f"GapGenerator - API type: {openai.api_type}")
        self.logger.debug(f"GapGenerator - API base: {openai.api_base}")
        self.logger.debug(f"GapGenerator - API version: {openai.api_version}")
        self.logger.debug(f"GapGenerator - API key (first 5 chars): {openai.api_key[:5] if openai.api_key else 'None'}")
        self.logger.debug(f"GapGenerator - Deployment name: {os.getenv('AZURE_DEPLOYMENT_NAME')}")
        
        # If any configuration is missing, apply direct configuration
        if not all([openai.api_type, openai.api_base, openai.api_version, openai.api_key, os.getenv('AZURE_DEPLOYMENT_NAME')]):
            self.logger.warning("OpenAI configuration incomplete. Applying direct configuration.")
            setup_openai_config()
            self.logger.debug(f"After direct config - API type: {openai.api_type}")
            self.logger.debug(f"After direct config - API base: {openai.api_base}")
            self.logger.debug(f"After direct config - API version: {openai.api_version}")
            self.logger.debug(f"After direct config - API key (first 5 chars): {openai.api_key[:5] if openai.api_key else 'None'}")
            self.logger.debug(f"After direct config - Deployment name: {os.getenv('AZURE_DEPLOYMENT_NAME')}")
        
    def setup_styles(self):
        """Setup Excel styling for the output worksheet"""
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_font = Font(bold=True)
        self.wrap_text_alignment = Alignment(wrap_text=True, vertical='top')
        
    def apply_excel_formatting(self, cell, value):
        """Apply proper formatting to Excel cell with special handling for line breaks"""
        # Ensure line breaks are preserved in Excel
        if isinstance(value, str) and ('\n' in value or '\r\n' in value):
            # Replace all line breaks with Excel's line feed character (ASCII 10)
            formatted_value = value.replace('\r\n', '\n')
            cell.value = formatted_value
        else:
            cell.value = value
            
        # Apply text wrapping
        cell.alignment = self.wrap_text_alignment
        
    def process_template(self, input_file: str) -> openpyxl.Workbook:
        """
        Process the input template and generate gaps and recommendations.
        
        Args:
            input_file: Path to the input Excel file
            
        Returns:
            The processed workbook with gaps and recommendations
        """
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(input_file)
            sheet = workbook.active
            
            # Find the maximum row with data
            max_row = sheet.max_row
            
            # Detect headers
            headers = {}
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers[cell_value.strip()] = col
            
            # Check if required columns exist
            required_columns = [
                'Control ID', 'Control Name', 'Control Description', 
                'Application', 'Current Process', 'Gap Status',
                'Gap Title', 'Gap Description', 'Recommendation'
            ]
            
            for column in required_columns:
                if column not in headers:
                    raise ValueError(f"Required column '{column}' not found in the template")
            
            # Process each row in the template
            for row in range(2, max_row + 1):
                # Check if this is an empty row
                if not sheet.cell(row=row, column=headers['Control ID']).value:
                    continue
                    
                # Extract control information
                control_id = sheet.cell(row=row, column=headers['Control ID']).value
                control_name = sheet.cell(row=row, column=headers['Control Name']).value
                control_description = sheet.cell(row=row, column=headers['Control Description']).value
                application = sheet.cell(row=row, column=headers['Application']).value
                current_process = sheet.cell(row=row, column=headers['Current Process']).value
                gap_status = sheet.cell(row=row, column=headers['Gap Status']).value
                
                # Skip if gap status is empty or "No Gap"
                if not gap_status or gap_status.lower() == "no gap":
                    continue
                
                # Check if Gap Title, Gap Description, and Recommendation are already filled
                gap_title_cell = sheet.cell(row=row, column=headers['Gap Title'])
                gap_desc_cell = sheet.cell(row=row, column=headers['Gap Description'])
                recommendation_cell = sheet.cell(row=row, column=headers['Recommendation'])
                
                # Skip if all gap fields are already filled
                if (gap_title_cell.value and gap_desc_cell.value and recommendation_cell.value):
                    continue
                
                # Generate gaps and recommendations
                gap_title, gap_description, recommendation = self.generate_gap_content(
                    control_id=control_id,
                    control_name=control_name,
                    control_description=control_description,
                    application=application,
                    current_process=current_process,
                    gap_status=gap_status
                )
                
                # Update the cells with generated content
                self.apply_excel_formatting(gap_title_cell, gap_title)
                self.apply_excel_formatting(gap_desc_cell, gap_description)
                self.apply_excel_formatting(recommendation_cell, recommendation)
                
            # Auto-adjust column widths for better readability
            for col in range(1, sheet.max_column + 1):
                max_length = 0
                column = openpyxl.utils.get_column_letter(col)
                for row in range(1, max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = max(max_length, 10)
                sheet.column_dimensions[column].width = min(adjusted_width, 50)
            
            return workbook
            
        except Exception as e:
            self.logger.error(f"Error processing template: {str(e)}")
            raise
    
    def generate_gap_content(
        self, 
        control_id: str, 
        control_name: str, 
        control_description: str,
        application: str,
        current_process: str,
        gap_status: str
    ) -> Tuple[str, str, str]:
        """
        Generate gap title, description, and recommendation using Azure OpenAI.
        
        Args:
            control_id: The control ID
            control_name: The name of the control
            control_description: Description of the control
            application: The application name
            current_process: Description of the current process
            gap_status: The gap status (Not Implemented, Partially Implemented, Improvement Needed)
            
        Returns:
            Tuple containing gap title, gap description, and recommendation
        """
        # Create a cache key
        cache_key = f"{control_id}_{application}_{gap_status}_{hash(current_process)}"
        
        # Check if response is in cache
        if cache_key in self.cache:
            return self.cache[cache_key]
            
        # Extract the base control ID (e.g., "APD-01" from "APD-01-Azure")
        base_control_id = control_id.split('-')
        if len(base_control_id) >= 2:
            base_control_id = f"{base_control_id[0]}-{base_control_id[1]}"
        else:
            base_control_id = control_id
            
        # Normalize gap status to match examples in gap_definitions.py
        normalized_gap_status = gap_status.lower()
        if "informal" in normalized_gap_status or "partially" in normalized_gap_status:
            normalized_gap_status = "informal process"
        else:
            normalized_gap_status = "gap"
            
        self.logger.debug(f"Looking for examples with base control ID: {base_control_id} and status: {normalized_gap_status}")
            
        # Find examples for this control and gap status
        examples = find_gap_examples_by_control_and_status(base_control_id, normalized_gap_status)
        
        # If no exact match, get all examples for this control
        if not examples:
            self.logger.debug(f"No exact status match, getting all examples for control: {base_control_id}")
            examples = find_gap_examples_by_control(base_control_id)
            
        # Generate example text for the prompt
        example_text = ""
        if examples:
            for i, example in enumerate(examples, 1):
                example_text += f"Example {i}:\n"
                example_text += f"Gap Title: {example.gap_title}\n"
                example_text += f"Gap Description: {example.gap_description}\n"
                example_text += f"Recommendation: {example.recommendation}\n\n"
        else:
            self.logger.warning(f"No examples found for control ID: {base_control_id}")
        
        # Build the prompt for OpenAI
        prompt = f"""
I need to generate a Gap Title, Gap Description, and Recommendation for an IT General Control (ITGC) that has a gap.

Control Information:
- Control ID: {control_id}
- Control Name: {control_name}
- Control Description: {control_description}
- Application: {application}
- Current Process: {current_process}
- Gap Status: {gap_status}

Based on the Control Information, please generate a concise Gap Title, a detailed Gap Description, and a specific Recommendation.

IMPORTANT: Your response MUST closely match the style, format, language, and length of the examples provided. DO NOT use asterisks (***) in your response.

The Gap Title should follow this format: "Absence of Formalized Process for [Application] [Control Name]" or "Informal [Control Name] Process for [Application]", depending on the gap status.

The Gap Description should be structured in paragraphs like the examples, explaining what specifically is missing or inadequate in the current process compared to the control requirements. Reference specifics from the Current Process description.

The Recommendation should begin with "For the system(s) listed in column E, perform the following steps: " followed by an empty line, then numbered steps (1, 2, 3, etc.) that are specific and actionable. Include a blank line between each numbered step.

For example, format the recommendation EXACTLY like this:

For the system(s) listed in column E, perform the following steps:

1. First recommendation step.
2. Second recommendation step.
3. Third recommendation step.

{example_text}

Please format your response in the following structure:
Gap Title: [Your generated gap title]
Gap Description: [Your generated gap description]
Recommendation: [Your generated recommendation]

Your response should be tailored to match the style and format of the examples while being specific to this control and application.
"""
        
        try:
            # Get deployment name from environment, with fallback
            deployment_name = os.getenv('AZURE_DEPLOYMENT_NAME')
            if not deployment_name:
                self.logger.warning("AZURE_DEPLOYMENT_NAME not found in environment variables, using default 'gpt-4o-it-risk'")
                deployment_name = 'gpt-4o-it-risk'
                
            self.logger.debug(f"Using deployment name: {deployment_name}")
                
            # Call Azure OpenAI API
            response = openai.ChatCompletion.create(
                engine=deployment_name,
                messages=[
                    {"role": "system", "content": "You are an IT audit expert who specializes in IT General Controls (ITGC). You provide detailed gap assessments and remediation recommendations. Follow the examples exactly."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,  # Lower temperature for more consistent results
                max_tokens=1200
            )
            
            response_text = response.choices[0].message['content'].strip()
            
            # Remove any asterisks from the response
            response_text = response_text.replace("***", "").replace("**", "").replace("*", "")
            
            # Parse the response to extract the three components
            gap_title = ""
            gap_description = ""
            recommendation = ""
            
            current_section = None
            for line in response_text.split('\n'):
                if line.startswith("Gap Title:"):
                    current_section = "title"
                    gap_title = line[len("Gap Title:"):].strip()
                elif line.startswith("Gap Description:"):
                    current_section = "description"
                    gap_description = line[len("Gap Description:"):].strip()
                elif line.startswith("Recommendation:"):
                    current_section = "recommendation"
                    recommendation = line[len("Recommendation:"):].strip()
                elif current_section == "title" and not line.startswith("Gap Description:"):
                    gap_title += " " + line.strip()
                elif current_section == "description" and not line.startswith("Recommendation:"):
                    gap_description += " " + line.strip() if gap_description else line.strip()
                elif current_section == "recommendation":
                    recommendation += " " + line.strip() if recommendation else line.strip()
            
            # If we have examples, try to match the format better
            if examples and (not gap_title or not gap_description or not recommendation):
                example = examples[0]
                
                # Use example as template if needed
                if not gap_title:
                    gap_title = example.gap_title.replace(example.application or "", application)
                
                if not gap_description:
                    gap_description = example.gap_description.replace(example.application or "", application)
                    
                if not recommendation:
                    recommendation = example.recommendation
            
            # Verify we have all three parts
            if not (gap_title and gap_description and recommendation):
                # Fallback: Split the response into three roughly equal parts
                parts = response_text.split('\n\n', 2)
                if len(parts) >= 3:
                    gap_title = parts[0].replace("Gap Title:", "").strip()
                    gap_description = parts[1].replace("Gap Description:", "").strip()
                    recommendation = parts[2].replace("Recommendation:", "").strip()
                else:
                    # Last resort fallback
                    self.logger.warning("Unable to properly parse OpenAI response, using fallback content.")
                    if examples:
                        example = examples[0]
                        gap_title = example.gap_title.replace(example.application or "", application)
                        gap_description = example.gap_description.replace(example.application or "", application)
                        recommendation = example.recommendation
                    else:
                        gap_title = f"Absence of Formalized Process for {application} {control_name}"
                        gap_description = f"The current implementation of {control_name} for {application} does not meet the requirements specified in the control description. The current process '{current_process}' has a {gap_status.lower()} status."
                        recommendation = self.format_recommendation(f"For the system(s) listed in column E, perform the following steps:\r\n\r\n1. Develop a formally documented {control_name} policy for {application}.\r\n2. Implement a standardized process with proper documentation and approvals.\r\n3. Establish formal reviews to ensure ongoing compliance with the control requirements.")
            
            # Format the recommendation to ensure proper line breaks between numbered steps
            recommendation = self.format_recommendation(recommendation)
            
            # Final sanity check to ensure the format is correct
            if "For the system(s) listed in column E, perform the following steps:" in recommendation:
                # Get intro and steps parts
                intro_phrase = "For the system(s) listed in column E, perform the following steps:"
                parts = recommendation.split(intro_phrase, 1)
                if len(parts) == 2:
                    steps = parts[1].strip()
                    
                    # Add a single blank line after the intro
                    formatted_rec = f"{intro_phrase}\n\n"  # Use \n for Excel compatibility
                    
                    # Process each line, ensuring each numbered step is on its own line
                    import re
                    
                    # First, make sure each number starts on a new line
                    for i in range(1, 20):  # Handle steps 1-19
                        steps = re.sub(f"([^\n]){i}\\.", f"\n{i}.", steps)
                    
                    # Split the steps by line breaks
                    step_lines = re.split(r'\r\n|\n', steps)
                    for line in step_lines:
                        line = line.strip()
                        if line and re.match(r'^\d+\.', line):
                            # Only add numbered lines
                            formatted_rec += f"{line}\n"
                    
                    # Use the reformatted recommendation - with \n line breaks for Excel
                    recommendation = formatted_rec
            
            # Cache the result
            result = (gap_title, gap_description, recommendation)
            self.cache[cache_key] = result
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error generating gap content: {str(e)}")
            
            # Fallback content in case the API call fails
            if examples:
                example = examples[0]
                return (
                    example.gap_title.replace(example.application or "", application),
                    example.gap_description.replace(example.application or "", application),
                    self.format_recommendation(example.recommendation)
                )
            else:
                return (
                    f"Absence of Formalized Process for {application} {control_name}",
                    f"The current implementation of {control_name} for {application} does not meet the requirements specified in the control description. The current process '{current_process}' has a {gap_status.lower()} status.",
                    self.format_recommendation(f"For the system(s) listed in column E, perform the following steps:\r\n\r\n1. Develop a formally documented {control_name} policy for {application}.\r\n2. Implement a standardized process with proper documentation and approvals.\r\n3. Establish formal reviews to ensure ongoing compliance with the control requirements.")
                )
                
    def format_recommendation(self, recommendation: str) -> str:
        """
        Format the recommendation to ensure proper line breaks between numbered steps and 
        an empty line after the intro phrase.
        
        Args:
            recommendation: The recommendation text to format
            
        Returns:
            Properly formatted recommendation
        """
        try:
            # The intro phrase we're looking for
            intro_phrase = "For the system(s) listed in column E, perform the following steps:"
            
            # If the phrase isn't in the recommendation, return it unchanged
            if intro_phrase not in recommendation:
                return recommendation
                
            # Split by the intro phrase
            parts = recommendation.split(intro_phrase, 1)
            if len(parts) != 2:
                return recommendation
                
            steps_part = parts[1].strip()
            
            # Start with a clean formatted recommendation
            formatted_rec = f"{intro_phrase}\n\n"  # One blank line after intro
            
            # Simplest and most reliable approach: 
            # Use regex to handle each numbered step
            import re
            
            # First normalize line breaks
            steps_part = steps_part.replace('\r\n', '\n')
            
            # Add line breaks before numbers if they don't already have them
            for i in range(1, 20):  # Handle up to 19 steps
                # If a number isn't preceded by a line break, add one
                steps_part = re.sub(f'([^\n]){i}\\.', f'\\1\n{i}.', steps_part)
            
            # Make sure first step has a line break if needed
            if steps_part and not steps_part.startswith('\n'):
                steps_part = '\n' + steps_part
                
            # Extract each numbered step
            numbered_steps = []
            lines = steps_part.split('\n')
            for line in lines:
                line = line.strip()
                if line and re.match(r'^\d+\.', line):
                    numbered_steps.append(line)
            
            # Add each step on its own line
            for step in numbered_steps:
                formatted_rec += f"{step}\n"
                
            return formatted_rec
            
        except Exception as e:
            self.logger.error(f"Error formatting recommendation: {str(e)}")
            
            # As a last resort fallback, use a very simple regex approach
            try:
                # Split by intro phrase
                intro_phrase = "For the system(s) listed in column E, perform the following steps:"
                if intro_phrase not in recommendation:
                    return recommendation
                    
                parts = recommendation.split(intro_phrase, 1)
                steps = parts[1] if len(parts) > 1 else ""
                
                # Simple regex to ensure each number starts on a new line
                import re
                formatted_steps = ""
                for i in range(1, 20):  # Handle up to 19 steps
                    pattern = f"{i}\\.(.*?)(?={i+1}\\.|$)"
                    matches = re.findall(pattern, steps, re.DOTALL)
                    if matches:
                        for match in matches:
                            formatted_steps += f"{i}. {match.strip()}\n"
                
                return f"{intro_phrase}\n\n{formatted_steps}"
                
            except Exception as inner_e:
                self.logger.error(f"Error in fallback formatting: {str(inner_e)}")
                # Very last resort: just ensure there's a line break after each period
                return recommendation.replace('. ', '.\n') 