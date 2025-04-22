import logging
import os
import openai
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from models.control_definitions import Control, STANDARD_CONTROLS
from typing import List
from openpyxl.worksheet.datavalidation import DataValidation

# Define the headers we care about from the scoping document
RELEVANT_HEADERS = {
    "System Name",
    "System Administration Responsibility",
    "Access Provisioning Process",
    "Access Removal Process",
    "Role Modification Capability",
    "Role Review",
    "System Accounts",
    "Credential Storage for System Accounts",
    "System Account Credential Access",
    "User Access Review",
    "Activity Logging Functionality",
    "Admin Activity Review",
    "User Authentication Method",
    "Authentication Configuration Review",
    "Change Management Process",
    "Separate Environments",
    "Change Review Process",
    "Automated Jobs Overview",
    "Job Management Tools",
    "Job Failure Resolution",
    "Backup Frequency",
    "Backup Types",
    "Backup Failure Resolution",
    "SOC report Review",
    "Segregation of Duties",
    "Change Access"
}

class MatrixGenerator:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.setup_styles()
        self.setup_openai()
        self.description_cache = {}  # Simple cache for API responses
        
    def setup_styles(self):
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def setup_openai(self):
        """Set up OpenAI with proper configuration"""
        # OpenAI configuration is now handled in app.py
        pass

    def generate_control_description(self, control_template: Control, system_info: dict) -> str:
        """Use GPT to generate a customized control description based on both template and scoping info."""
        
        # Create a cache key based on control ID AND short name (to differentiate variants) and system name
        cache_key = f"{control_template.control_id}_{control_template.short_name}_{system_info['name']}"
        
        # Check if we have a cached response
        if cache_key in self.description_cache:
            return self.description_cache[cache_key]

        # Extract relevant scoping information based on control headers
        relevant_scoping_info = []
        for header in control_template.scoping_headers:
            header_key = header.lower().replace(" ", "_")
            if header_key in system_info:
                relevant_scoping_info.append(f"{header}: {system_info[header_key]}")

        # Add system administration responsibility if the template contains [INSERT ADMIN]
        admin_info = ""
        if "[INSERT ADMIN]" in control_template.description:
            admin_key = "system_administration_responsibility"
            if admin_key in system_info and system_info[admin_key]:
                admin_info = f"\nSystem Administration Responsibility: {system_info[admin_key]}"

        prompt = f"""
        Using the following information, generate a control description that matches the style and format of the base description while incorporating the actual system details and processes.

        Control Template:
        {control_template.description}

        Control Short Name:
        {control_template.short_name}

        Control Type:
        {control_template.control_type}

        System Information:
        System Name: {system_info['name']}{admin_info}
        Relevant Scoping Details:
        {chr(10).join(relevant_scoping_info)}

        Requirements:
        1. Use only factual, observable details from the provided scoping information
        2. Replace placeholder text as follows:
        - Replace "[INSERT SYSTEM NAME]" with the actual system name
        - Replace "[INSERT ADMIN]" with the system administration responsibility team/role
        - Replace other placeholders with appropriate values from scoping details
        3. Match the structure, tone, and formality of the base description EXACTLY
        4. Include ONLY tools, processes, and roles that are explicitly mentioned in the scoping information
        5. Keep to the same length as the base description
        6. IMPORTANT: Ensure the description maintains the core process described in the control's short name (e.g., if it's "Manual Approval", make sure the process is manual)
        7. If the scoping information contradicts the template, use the scoping information but maintain the key characteristic (automated vs manual, etc.)
        8. DO NOT add explanatory details or examples that aren't in the base template
        9. DO NOT elaborate or expand on processes beyond what's in the template
        10. Return ONLY the description text, without any metadata or formatting
        11. CRITICAL: NEVER use personal names in the description. Instead, use:
           - Job titles (e.g., "IT Manager", "System Administrator")
           - Team names (e.g., "IT team", "Security team")
           - Role-based terms (e.g., "the administrator", "the reviewer")
           - Generic terms (e.g., "management", "leadership")

        IMPORTANT:
        - Keep the same sentence structure as the template
        - Do not add explanatory clauses or examples
        - Do not elaborate on processes or add implementation details
        - Replace placeholders only with direct, equivalent information
        - Maintain the exact same level of detail as the template
        - For [INSERT ADMIN], use the exact team/role specified in System Administration Responsibility
        - Maintain the core characteristics indicated in the control's short name (e.g., "Manual" vs "Automated")
        - NEVER include personal names - use job titles, team names, or role-based terms instead

        Return only the customized control description text, with no additional formatting or labels.
        """

        try:
            response = openai.ChatCompletion.create(
                engine=os.getenv('AZURE_DEPLOYMENT_NAME', 'gpt-4o-it-risk'),
                messages=[
                    {"role": "system", "content": "You are a technical writer. Write clear, factual control descriptions without elaboration or additional details. Match the template exactly, only replacing placeholder values."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=750
            )
            description = response.choices[0].message['content'].strip()
            
            # Cache the result before returning
            self.description_cache[cache_key] = description
            return description
        except Exception as e:
            self.logger.error(f"Error generating control description: {str(e)}")
            # Fallback to basic replacement if GPT fails
            description = control_template.description.replace("[INSERT SYSTEM NAME]", system_info['name'])
            if "[INSERT ADMIN]" in description:
                admin = system_info.get('system_administration_responsibility', 'System Administrator')
                description = description.replace("[INSERT ADMIN]", admin)
            
            # Add variant indicator for fallback mode to ensure distinction between variants
            if "Manual" in control_template.short_name and "Manual" not in description:
                description = description.replace(".", ". (Manual process).")
            elif "Automated" in control_template.short_name and "Automated" not in description:
                description = description.replace(".", ". (Automated process).")
                
            return description

    def _format_system_details(self, system_info: dict) -> str:
        """Format system details for the prompt."""
        details = []
        for key, value in system_info.items():
            if key != 'name':  # name is already included in the main prompt
                details.append(f"{key.replace('_', ' ').title()}: {value}")
        return "\n".join(details)

    def process_scoping_document(self, input_file_path):
        """Process the scoping document and generate the matrix."""
        try:
            # Read systems and their existing controls
            systems = self._read_scoping_document(input_file_path)
            
            # Generate matrix
            matrix_wb = self._generate_matrix(systems)
            
            return matrix_wb
        except Exception as e:
            self.logger.error(f"Error processing scoping document: {str(e)}")
            raise

    def _read_scoping_document(self, file_path: str) -> list:
        """Read and process the scoping document, focusing only on relevant headers."""
        try:
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active
            
            # Get all headers from first row
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            
            # Verify first column is "System Name"
            if headers[0].lower() != "system name":
                raise ValueError("First column must be 'System Name'")
            
            # Find indices of relevant headers
            header_indices = {'Application Name': 0}  # We know System Name is first column
            for idx, header in enumerate(headers):
                # Check if header exists in RELEVANT_HEADERS (case-insensitive)
                for relevant_header in RELEVANT_HEADERS:
                    if header.lower() == relevant_header.lower():
                        header_indices[relevant_header] = idx
                        break
            
            # Process each row until we hit an empty system name
            systems = []
            for row in list(ws.rows)[1:]:  # Skip header row
                system_name = row[0].value  # Get value from first column
                
                # Stop if we hit a row with no system name
                if not system_name:
                    break
                    
                # Create app info dictionary
                app_info = {
                    "name": system_name
                }
                
                # Only process columns with headers we care about
                for header, idx in header_indices.items():
                    if header != "Application Name":  # Skip system name as we already got it
                        key = header.lower().replace(" ", "_")
                        app_info[key] = row[idx].value or ""
                
                systems.append(app_info)
            
            self.logger.info(f"Processed {len(systems)} systems from scoping document")
            return systems
            
        except Exception as e:
            self.logger.error(f"Error reading scoping document: {str(e)}")
            raise

    def select_appropriate_control(self, control_variants: List[Control], system_info: dict) -> tuple[Control, bool]:
        """Use GPT to select the most appropriate control variant based on scoping information.
        Returns a tuple of (selected_control, is_na)"""
        
        # First check for explicit N/A responses or empty values
        for variant in control_variants:
            for header in variant.scoping_headers:
                # Split multiple headers (e.g., "User Activity Review; Activity Logging Functionality")
                sub_headers = [h.strip() for h in header.split(';')]
                
                for sub_header in sub_headers:
                    header_key = sub_header.lower().replace(" ", "_")
                    if header_key in system_info:
                        value = str(system_info[header_key]).strip().upper()
                        # Expanded N/A check to include variations and negative responses
                        if (value.startswith("N/A") or 
                            not value or 
                            value == "NO" or 
                            value == "NONE" or 
                            value == "NOT APPLICABLE" or
                            value == "NOT PERFORMED" or
                            value == "NOT REQUIRED"):
                            return control_variants[0], True  # Return first variant and mark as N/A

        # If only one variant, no need to select
        if len(control_variants) == 1:
            return control_variants[0], False

        # Build detailed scoping information for evaluation
        relevant_info = ""
        for header in set().union(*[set(variant.scoping_headers) for variant in control_variants]):
            header_key = header.lower().replace(" ", "_")
            if header_key in system_info:
                relevant_info += f"{header}: {system_info[header_key]}\n"

        variant_descriptions = []
        for i, ctrl in enumerate(control_variants):
            description = f"Option {i+1}:\n"
            description += f"Name: {ctrl.short_name}\n"
            description += f"Description: {ctrl.description}\n"
            description += f"Evaluation Criteria: {ctrl.evaluation_criteria}\n"
            description += f"Required Headers: {', '.join(ctrl.scoping_headers)}"
            variant_descriptions.append(description)

        prompt = f"""
        Analyze the following scoping information and select the most appropriate control variant.
        You must be very strict in your evaluation - if the scoping information doesn't FULLY satisfy 
        the evaluation criteria, consider it a non-match.
        
        Scoping Information:
        {relevant_info}

        Control Variants:
        {'\n\n'.join(variant_descriptions)}

        Requirements:
        1. Compare the scoping information against each control's evaluation criteria
        2. Select the control that best matches the described process
        3. For example, if scoping information mentions automated workflows, select the automated variant
        4. If scoping information mentions manual processes or approvals, select the manual variant
        5. Match the process described in scoping information to the appropriate control variant
        6. Return ONLY the number of the selected option (1, 2, etc.)
        """

        try:
            response = openai.ChatCompletion.create(
                engine=os.getenv('AZURE_DEPLOYMENT_NAME', 'gpt-4o-it-risk'),
                messages=[
                    {"role": "system", "content": "You are a control selection expert. Analyze information and select the most appropriate control variant."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=10
            )
            
            response_text = response.choices[0].message['content'].strip()
            
            # Try to parse as control selection
            try:
                # Extract just the digit if there's other text
                import re
                digits = re.findall(r'\d+', response_text)
                if digits:
                    selection = int(digits[0]) - 1
                    if 0 <= selection < len(control_variants):
                        self.logger.info(f"Selected variant {selection+1} for control {control_variants[0].control_id}")
                        return control_variants[selection], False
            except ValueError:
                pass
            
            # Default to first variant if response is unclear
            return control_variants[0], False
        
        except Exception as e:
            self.logger.error(f"Error selecting control variant: {str(e)}")
            return control_variants[0], False  # Default to first variant on error

    def _generate_matrix(self, systems):
        wb = Workbook()
        ws = wb.active
        ws.title = "Risk and Controls Matrix"

        headers = [
            "Risk ID",
            "Risk Description",
            "Control ID",
            "Control Name",
            "Control Description",
            "Application",
            "Key / Non-Key",
            "Manual / IT-dependent manual / Automated",
            "Preventative / Detective",
            "Frequency",
            "Current Process",
            "Gap Status",
            "Gap Title",
            "Gap Description",
            "Recommendation"
        ]

        # Create data validation for Gap Status column (column 12)
        gap_status_dv = DataValidation(type="list", formula1='"No Gap,Gap,Informal Process"', allow_blank=True)
        ws.add_data_validation(gap_status_dv)
        
        # Apply data validation to Gap Status column in all rows we might use
        # Excel has a maximum of 1048576 rows, but we'll use a more reasonable number
        for i in range(2, 5000):  # Start from row 2 (after headers) to row 5000
            gap_status_cell = f"L{i}"  # Column L is the 12th column (Gap Status)
            gap_status_dv.add(gap_status_cell)

        self._setup_headers(ws, headers)
        current_row = 2

        # Group controls by their base ID (e.g., "APD-01" for both "APD-01" and "APD-01B")
        control_groups = {}
        for control in STANDARD_CONTROLS:
            base_id = control.control_id.split('-')[0] + '-' + control.control_id.split('-')[1]
            if base_id not in control_groups:
                control_groups[base_id] = []
            control_groups[base_id].append(control)

        # Process each system separately
        for system in systems:
            # Add system name as a section header
            cell = ws.cell(row=current_row, column=1, value=f"System: {system['name']}")
            cell.font = Font(bold=True)
            current_row += 1
            
            # Check if this system has any flags in its scoping information
            system_has_flags = False
            for key, value in system.items():
                if value and isinstance(value, str) and "*Flag" in value:
                    self.logger.info(f"System {system['name']} has flag in field {key}: {value[:50]}...")
                    system_has_flags = True
                    break

            # Process each control group
            for control_variants in control_groups.values():
                # Select the appropriate variant but keep all variants
                selected_control, is_na = self.select_appropriate_control(control_variants, system)
                
                # Find out which scoping headers are relevant for this control
                relevant_headers = []
                for variant in control_variants:
                    # Add each scoping header, handling both string and list formats
                    for header_item in variant.scoping_headers:
                        if header_item not in relevant_headers:
                            relevant_headers.append(header_item)
                
                # Check if any relevant scoping fields contain flags
                control_has_flags = False
                for header in relevant_headers:
                    # Convert header to a dictionary key format consistently
                    header_key = header.lower().replace(" ", "_")
                    if header_key in system and isinstance(system[header_key], str) and "*Flag" in system[header_key]:
                        self.logger.info(f"Flag found in {header_key} for control {control_variants[0].control_id}")
                        control_has_flags = True
                        break
                
                # Add all variants to the matrix
                for variant in control_variants:
                    customized_description = self.generate_control_description(variant, system)
                    
                    # If N/A, highlight all variants in yellow
                    # If not N/A, highlight selected variant in green and others in yellow
                    is_selected = not is_na and (variant == selected_control)
                    
                    # Check if description contains a flag
                    description_has_flag = "*Flag" in customized_description
                    
                    # Pass the flag status for this specific control
                    self._add_control_row(
                        ws, 
                        current_row, 
                        variant, 
                        customized_description, 
                        system['name'], 
                        is_selected, 
                        is_na,
                        system=system,  # Pass the entire system dictionary
                        has_flag=control_has_flags or description_has_flag  # Pass explicit flag status
                    )
                    current_row += 1
            
            # Add a blank row between systems
            current_row += 1

        return wb

    def _setup_headers(self, ws, headers):
        """Set up and format the headers row"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
            ws.column_dimensions[chr(64 + col)].width = 20

    def _add_control_row(self, ws, row, control, description, system_name, is_selected, is_na, system=None, has_flag=False):
        """Add a control row with formatting"""
        try:
            # Create system-specific control ID
            name_suffix = str(system_name).replace(" ", "") if system_name else "Unknown"
            control_id = f"{control.control_id}-{name_suffix}"
            
            # Get the current process information from the system scoping details
            current_process = ""
            
            # Log the control and its scoping headers for debugging
            self.logger.info(f"Processing control {control_id} with scoping headers: {control.scoping_headers}")
            
            # Collect all the relevant scoping information
            scoping_info = []
            if system:  # Check if system info was provided
                # For logging, show all available system keys
                system_keys = list(system.keys())
                self.logger.info(f"Available system keys for {system_name}: {system_keys}")
                
                # Process each header in the list of headers for this control
                for header_item in control.scoping_headers:
                    self.logger.info(f"Processing header item: {header_item}")
                    
                    # Convert header to a dictionary key format
                    header_key = header_item.lower().replace(" ", "_")
                    
                    # Check if this key exists in the system dictionary
                    if header_key in system:
                        value = system[header_key]
                        if value:  # Only add if there's a value
                            self.logger.info(f"Found value for {header_key}: {value[:50]}...")
                            scoping_info.append(f"{header_item}: {value}")
                        else:
                            self.logger.info(f"Empty value for {header_key}")
                    else:
                        self.logger.info(f"Header key '{header_key}' not found in system info for {system_name}")
                        
                        # Check if the header might be formatted differently in the system dictionary
                        potential_matches = [k for k in system.keys() if header_key in k or k in header_key]
                        if potential_matches:
                            self.logger.info(f"Potential matches for {header_key}: {potential_matches}")
            
            # Join with periods between items if there are any
            if scoping_info:
                self.logger.info(f"Final scoping info for {control_id}: {scoping_info}")
                current_process = ". ".join(scoping_info)
            else:
                self.logger.info(f"No scoping info found for {control_id}")
            
            data = [
                control.risk_id,  # Risk ID
                control.risk_description,  # Risk Description
                control_id,
                control.short_name,
                description,
                system_name,  # Application column
                "Key" if control.key_control else "Non-key",
                control.control_type,
                control.nature,
                control.frequency,
                current_process,  # Current Process column
                "",  # Gap Status (will be empty with data validation)
                "",  # Gap Title
                "",  # Gap Description
                ""   # Recommendation
            ]

            # Also check if any cell in the data contains "*Flag" as a backup detection
            if not has_flag:
                for value in data:
                    value_str = str(value).strip() if value is not None else ""
                    if "*Flag" in value_str:
                        has_flag = True
                        self.logger.info(f"Flag detected in row data for control {control_id}")
                        break
            
            # Determine fill color:
            # Red for flagged items
            # Yellow for N/A or non-selected
            # Green for selected (and not flagged)
            if has_flag:
                fill_color = "FF6B6B"  # Red color for flagged items
            else:
                fill_color = "FFD700" if is_na else ("90EE90" if is_selected else "FFD700")  # Yellow for N/A or non-selected, Green for selected
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Apply fill color
                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )
                
                # Adjust column width based on content
                if len(str(value)) > ws.column_dimensions[chr(64 + col)].width:
                    ws.column_dimensions[chr(64 + col)].width = min(len(str(value)), 50)
        except Exception as e:
            self.logger.error(f"Error in _add_control_row: {str(e)}")
            raise