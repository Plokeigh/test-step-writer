import os
import re
import sys
import pandas as pd
import openpyxl
from copy import copy # Import copy for styles
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define color constants
RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
GREY_FILL = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
DARK_GREEN_FILL = PatternFill(start_color="FF006400", end_color="FF006400", fill_type="solid")
HEADER_FONT = Font(color="FF000000", bold=True, name='Arial', sz=10)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))

def extract_system_names(control_ids):
    """Extract unique system names from control IDs."""
    system_names = set()
    pattern = r'[A-Z]+-\d+-(.+)'
    
    for control_id in control_ids:
        if not isinstance(control_id, str):
            continue
            
        match = re.match(pattern, control_id)
        if match:
            system_name = match.group(1)
            system_names.add(system_name)
    
    return sorted(list(system_names))

def process_rcm_file(input_file, template_file):
    """
    Process an RCM file and map controls to the high-level view template.
    
    Args:
        input_file: Path to the input RCM Excel file
        template_file: Path to the template Excel file
    
    Returns:
        Path to the output file
    """
    try:
        # Read the input RCM file
        df = pd.read_excel(input_file)
        
        # Find the Control ID and Gap Status columns
        control_id_col = None
        gap_status_col = None
        
        for col in df.columns:
            if isinstance(col, str):
                if "control id" in col.lower():
                    control_id_col = col
                elif "gap status" in col.lower():
                    gap_status_col = col
        
        if not control_id_col or not gap_status_col:
            raise ValueError("Could not find 'Control ID' and/or 'Gap Status' columns in the RCM file")
        
        # Extract unique system names from Control IDs
        system_names = extract_system_names(df[control_id_col].tolist())
        
        if not system_names:
            raise ValueError("No system names found in Control ID column. Ensure they follow the format: '[Control-Code]-[System]'")
        
        # Create a dictionary to store controls by system
        system_controls = {}
        
        # Iterate through the RCM rows and organize by system
        for _, row in df.iterrows():
            control_id = row[control_id_col]
            gap_status = row[gap_status_col] if not pd.isna(row[gap_status_col]) else None
            
            if not isinstance(control_id, str):
                continue
                
            pattern = r'([A-Z]+-\d+)-(.+)'
            match = re.match(pattern, control_id)
            
            if match:
                control_code = match.group(1)
                system_name = match.group(2)
                
                if system_name not in system_controls:
                    system_controls[system_name] = {}
                
                system_controls[system_name][control_code] = gap_status
        
        # Load the template file
        workbook = openpyxl.load_workbook(template_file)
        sheet = workbook.active
        sheet.title = "RCM - High Level"
        
        # Find the row and column for mapping
        control_row = None
        start_col = None
        
        # Identify the control codes in column C and find row 8 for system placement
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=3).value == "Control ID":
                control_row = row
                break
        
        # Starting column H for systems
        start_col = 8  # Column H
        
        if not control_row:
            raise ValueError("Could not find 'Control ID' in column C of the template")
        
        # Place system names in row 8 starting from column H
        system_row = 8
        for idx, system in enumerate(system_names):
            col = start_col + idx
            cell = sheet.cell(row=system_row, column=col)
            cell.value = system
            cell.fill = GREY_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGNMENT
        
        # Map controls for each system
        for row in range(control_row + 1, sheet.max_row + 1):
            control_code = sheet.cell(row=row, column=3).value
            
            if not control_code:
                continue
                
            for idx, system in enumerate(system_names):
                col = start_col + idx
                cell = sheet.cell(row=row, column=col)
                
                # Check if the system has this control
                if system in system_controls and control_code in system_controls[system]:
                    gap_status = system_controls[system][control_code]
                    
                    # Apply conditional formatting based on Gap Status
                    if isinstance(gap_status, str):
                        gap_status_lower = gap_status.lower().strip()
                        if "gap" in gap_status_lower and "no gap" not in gap_status_lower:
                            cell.fill = RED_FILL
                        elif "no gap" in gap_status_lower:
                            cell.fill = GREEN_FILL
                        elif "informal process" in gap_status_lower:
                            cell.fill = YELLOW_FILL
                    else:
                        # System doesn't have explicit status for this control
                        cell.fill = GREY_FILL
                else:
                    # System doesn't have this control
                    cell.fill = GREY_FILL
        
        # Apply borders to the application columns from row 7 to 24
        border_start_row = 7
        border_end_row = 24
        border_end_col = start_col + len(system_names) - 1

        for row_idx in range(border_start_row, border_end_row + 1):
            for col_idx in range(start_col, border_end_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = THIN_BORDER

        # Merge cells in row 7 to span all system columns
        if len(system_names) > 1:
            start_cell_ref = f"{get_column_letter(start_col)}{7}"
            end_cell_ref = f"{get_column_letter(border_end_col)}{7}"
            sheet.merge_cells(f"{start_cell_ref}:{end_cell_ref}")
            # Reapply border and alignment to the merged cell H7 (since merging can remove formatting)
            merged_cell = sheet[start_cell_ref]
            merged_cell.border = THIN_BORDER
            merged_cell.alignment = CENTER_ALIGNMENT # Ensure merged header row is centered
            # Add a value to the merged cell if desired (e.g., "Applications")
            # merged_cell.value = "Applications" 
            # merged_cell.font = HEADER_FONT # Optional: Apply header font if adding text

        # Load input workbook with openpyxl to access styles
        input_workbook = openpyxl.load_workbook(input_file)
        input_sheet = input_workbook.active # Assuming data is in the active sheet

        # Create and populate the 'RCM - Granular' sheet with original data and formatting
        granular_sheet = workbook.create_sheet(title="RCM - Granular")
        
        # Copy cell values and styles
        for row in input_sheet.iter_rows():
            for cell in row:
                new_cell = granular_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        # Copy column dimensions
        for col_letter, dimension in input_sheet.column_dimensions.items():
            granular_sheet.column_dimensions[col_letter].width = dimension.width
        
        # Copy row dimensions
        for row_idx, dimension in input_sheet.row_dimensions.items():
             if dimension.height is not None: # Only copy if height is set
                granular_sheet.row_dimensions[row_idx].height = dimension.height
        
        # Note: Merged cells from input_sheet are not automatically copied here.
        # If merging is needed in granular_sheet, it would require separate logic.

        # Save the output file
        output_dir = os.path.dirname(input_file)
        output_file = os.path.join(output_dir, "RCM-High-Level-View.xlsx")
        workbook.save(output_file)
        
        return output_file
        
    except Exception as e:
        raise Exception(f"Error processing RCM file: {str(e)}")

def main():
    """Main function to process an RCM file."""
    if len(sys.argv) < 3:
        print("Usage: python main.py <input_rcm_file> <template_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    template_file = sys.argv[2]
    
    try:
        output_file = process_rcm_file(input_file, template_file)
        print(f"Successfully created high-level view: {output_file}")
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main() 