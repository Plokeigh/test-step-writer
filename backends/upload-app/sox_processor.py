import os
from openai import AzureOpenAI
from docx import Document
import logging
import tempfile
import traceback
from dotenv import load_dotenv
from typing import List, Dict
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook
import json

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Load environment variables
load_dotenv()

# Configure Azure OpenAI Client
api_key = os.getenv("OPENAI_API_KEY")
api_base = os.getenv("OPENAI_API_BASE")
api_version = os.getenv("OPENAI_API_VERSION", "2024-08-01-preview")
engine = os.getenv("OPENAI_ENGINE", "gpt-4o")

# Validate configuration
if not api_key:
    logger.error("OpenAI API key not configured.")
    raise ValueError("OpenAI API key not configured. Please set OPENAI_API_KEY environment variable.")

# Initialize Azure OpenAI client
client = AzureOpenAI(
    api_key=api_key,
    api_version=api_version,
    azure_endpoint=api_base
)

logger.debug(f"SOX Processor - API base: {api_base}")
logger.debug(f"SOX Processor - API version: {api_version}")
logger.debug(f"SOX Processor - Engine: {engine}")

# API configuration
API_CONFIG = {
    "max_tokens": 3000,
    "temperature": 0.1,
    "top_p": 0.95,
    "frequency_penalty": 0,
    "presence_penalty": 0
}

def make_openai_request(system_prompt: str, user_prompt: str, max_tokens: int = None) -> str:
    """Centralized OpenAI API request handler with error handling and logging."""
    config = API_CONFIG.copy()
    if max_tokens:
        config["max_tokens"] = max_tokens
        
    try:
        logger.debug(f"Making OpenAI request with {len(user_prompt)} character prompt")
        response = client.chat.completions.create(
            model=engine,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            **config
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logger.error(f"OpenAI API error: {str(e)}")
        logger.error(traceback.format_exc())
        return f"Error processing request: {str(e)}"

def parse_sox_controls_excel(file_path: str) -> List[Dict]:
    """Parse the uploaded Excel file with SOX control information."""
    logger.info(f"Parsing SOX controls Excel file: {file_path}")
    
    try:
        # Read the Excel file
        df = pd.read_excel(file_path)
        
        # Expected columns: Ref ID (A), Control Description (B), Testing Attributes (C), Design Attributes (D), Evidence of Control (E)
        expected_columns = ['Ref ID', 'Control Description', 'Testing Attributes', 'Design Attributes', 'Evidence of Control']
        
        # If columns don't match exactly, try to map them
        if not all(col in df.columns for col in expected_columns):
            # Map column positions (assuming A, B, C, D, E structure)
            df.columns = expected_columns[:len(df.columns)]
        
        controls = []
        for index, row in df.iterrows():
            # Convert all values to string and handle NaN/None values
            ref_id = str(row.get('Ref ID', '')).strip()
            control_desc = str(row.get('Control Description', '')).strip()
            testing_attrs = str(row.get('Testing Attributes', '')).strip()
            design_attrs = str(row.get('Design Attributes', '')).strip()
            evidence_ctrl = str(row.get('Evidence of Control', '')).strip()
            
            # Replace 'nan' with empty string (pandas NaN converts to 'nan' when cast to string)
            if ref_id.lower() == 'nan':
                ref_id = ''
            if control_desc.lower() == 'nan':
                control_desc = ''
            if testing_attrs.lower() == 'nan':
                testing_attrs = ''
            if design_attrs.lower() == 'nan':
                design_attrs = ''
            if evidence_ctrl.lower() == 'nan':
                evidence_ctrl = ''
                
            control = {
                'ref_id': ref_id,
                'control_description': control_desc,
                'testing_attributes': testing_attrs,
                'design_attributes': design_attrs,
                'evidence_of_control': evidence_ctrl
            }
            
            # Only add if we have meaningful data (ref_id and control_description)
            if control['ref_id'] and control['control_description']:
                logger.debug(f"Adding control: {control['ref_id']} - {control['control_description'][:50]}...")
                controls.append(control)
            else:
                logger.debug(f"Skipping row {index} - missing ref_id or control_description")
        
        logger.info(f"Parsed {len(controls)} controls from Excel file")
        return controls
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        raise

def generate_test_steps_from_control(control_data: Dict) -> Dict:
    """Generate test steps and attributes for a single control."""
    
    # Check if this is an N/A scenario (only Control Description available)
    testing_attrs = control_data.get('testing_attributes', '').strip().upper()
    design_attrs = control_data.get('design_attributes', '').strip().upper()
    evidence_ctrl = control_data.get('evidence_of_control', '').strip().upper()
    
    # Enhanced N/A detection - check for various forms of empty/N/A values
    is_na_scenario = (testing_attrs in ['N/A', 'NA', '', 'NAN', 'NULL'] or 
                     design_attrs in ['N/A', 'NA', '', 'NAN', 'NULL'] or 
                     evidence_ctrl in ['N/A', 'NA', '', 'NAN', 'NULL'])
    
    if is_na_scenario:
        # Enhanced creative prompt for N/A scenarios - extrapolate from Control Description
        system_prompt = """You are a SOX compliance specialist creating test steps for a control where only the Control Description is available.

You must CREATIVELY EXTRAPOLATE appropriate test steps and evidence based SOLELY on the Control Description provided.

## REQUIREMENTS FOR N/A SCENARIOS:

### 1. ANALYZE THE CONTROL DESCRIPTION DEEPLY:
- Identify WHO performs the control (third parties, management, departments)
- Identify WHAT is being controlled (reviews, reconciliations, approvals, monitoring)
- Identify HOW the control operates (frequency, process steps, documentation)
- Identify potential EVIDENCE that would logically exist for this control

### 2. EXTRAPOLATE LOGICAL EVIDENCE:
Based on the control description, determine what evidence would logically be created:
- **Review Controls**: Reports, review documentation, sign-offs, exception reports
- **Third-party Controls**: Third-party reports, management review of reports, response documentation
- **Reconciliation Controls**: Reconciliation workbooks, variance analysis, investigation documentation
- **Authorization Controls**: Approval documentation, authorization matrices, signed documents
- **Monitoring Controls**: Monitoring reports, exception handling documentation, follow-up evidence

### 3. CREATE REALISTIC "OBTAIN EVIDENCE" STEP:
The first step should list SPECIFIC evidence items that would logically exist:
- Name: "Obtain Evidence"
- Description: "For a sample [period], obtain the following evidence: A) [specific evidence 1], B) [specific evidence 2], C) [specific evidence 3], D) [specific evidence 4]"
- Be SPECIFIC about what evidence would exist (not generic)

### 4. GENERATE LOGICAL TEST STEPS:
Create 4-6 test steps that would logically test this control:
- Extract test procedures from the control description
- Focus on validating the KEY ELEMENTS mentioned in the description
- Create meaningful attribute names and descriptions

### 5. EXAMPLES OF CREATIVE EXTRAPOLATION:

**If control mentions "third-party reviews":**
- Evidence: Third-party reports, management review documentation
- Test Steps: "Inspect Third-party Reports", "Verify Management Review"

**If control mentions "discrepancy resolution":**
- Evidence: Discrepancy reports, resolution documentation
- Test Steps: "Inspect Discrepancy Identification", "Verify Resolution Process"

**If control mentions "adherence to guidelines":**
- Evidence: Guidelines documentation, adherence testing reports
- Test Steps: "Inspect Guidelines Compliance", "Verify Testing Procedures"

### 6. OUTPUT FORMAT:
```json
{
  "test_steps": [
    {
      "control_id": "[Use the Ref ID]",
      "name": "Test Step Name",
      "description": "Detailed description with specific references to control elements",
      "attribute_name": "Attribute Name or N/A",
      "attribute_description": "Past-tense description starting with 'Verified that...' or N/A"
    }
  ]
}
```

### 7. CRITICAL CREATIVITY GUIDELINES:
- READ the control description like a detective - what evidence MUST exist?
- Think about the BUSINESS PROCESS described - what documentation would be created?
- Consider COMPLIANCE REQUIREMENTS - what would auditors expect to see?
- Be SPECIFIC in evidence descriptions - avoid generic terms
- Create test steps that mirror the ACTUAL CONTROL ACTIVITIES described
- Ensure each test step validates a KEY ASPECT of the control described"""

        user_prompt = f"""CREATIVELY ANALYZE this SOX control and extrapolate realistic test steps and evidence:

Control ID: {control_data['ref_id']}
Control Description: {control_data['control_description']}

Based SOLELY on this control description, you must:
1. Identify what specific evidence would logically be created by this control process
2. List that evidence in the "Obtain Evidence" step as A), B), C), D) items
3. Create realistic test steps that would validate the key elements mentioned in the description
4. Generate meaningful attributes that test the control's effectiveness

Be as SPECIFIC and REALISTIC as possible - think like an experienced SOX auditor who understands what evidence would actually exist for this type of control."""

    else:
        # Enhanced detailed prompt for controls with full information
        system_prompt = """You are a SOX compliance specialist creating detailed test steps and test attributes.

## CRITICAL: DO NOT COPY TESTING ATTRIBUTES VERBATIM
You must TRANSFORM and REWRITE the testing attributes into proper test step format. 

### WHAT NOT TO DO (WRONG):
❌ Testing Attribute: "A) Monthly, the Investment Accounting Staff or Manager obtains statements and ownership percentages from the equity method sponsor or general partner."
❌ Wrong Test Step Description: "Monthly, the Investment Accounting Staff or Manager obtains statements and ownership percentages from the equity method sponsor or general partner."

### WHAT TO DO (CORRECT):
✅ Testing Attribute: "A) Monthly, the Investment Accounting Staff or Manager obtains statements and ownership percentages from the equity method sponsor or general partner."
✅ Correct Test Step Name: "Inspect Statement Procurement"
✅ Correct Test Step Description: "Inspect evidence that statements and ownership percentages were obtained from equity method sponsors or general partners for the selected month."
✅ Correct Attribute Name: "Statement Completeness"
✅ Correct Attribute Description: "Verified that all required statements and ownership percentages were obtained from appropriate equity method sponsors or general partners."

## TRANSFORMATION RULES:

### 1. FIRST TEST STEP - "Obtain Evidence" (MANDATORY)
- **Test Step Name**: "Obtain Evidence"
- **Test Step Description**: "For a sample month, obtain the following evidence of the [control purpose]: [List ALL items from Evidence of Control column as A), B), C), D) format]"
- **Attribute Name**: "N/A"
- **Attribute Description**: "N/A"

### 2. TRANSFORM EACH TESTING ATTRIBUTE (A, B, C, D, E):

For EACH lettered item in Testing Attributes, you must:

#### A. CREATE CONCISE TEST STEP NAME (2-4 words):
- Extract the CORE ACTION: "obtains statements" → "Inspect Statement Procurement"
- Extract the CORE ACTION: "calculates pro-rata" → "Inspect Calculation Process"  
- Extract the CORE ACTION: "reconciles the sum" → "Inspect Reconciliation"
- Extract the CORE ACTION: "reviews and approves" → "Inspect Review Evidence"
- Extract the CORE ACTION: "verifies completeness" → "Inspect IPE Verification"

#### B. REWRITE TEST STEP DESCRIPTION:
- Change from process description to testing instruction
- Use "Inspect", "Verify", "Review", "Validate" language
- Make it about what the TESTER does, not what the CONTROL PERFORMER does
- Keep it concise (1-2 sentences max)

#### C. CREATE MEANINGFUL ATTRIBUTE NAME (2-4 words):
- Focus on WHAT is being validated
- Examples: "Statement Completeness", "Calculation Accuracy", "Reconciliation Completeness", "Review Evidence", "IPE Verification"

#### D. WRITE ATTRIBUTE DESCRIPTION:
- Start with "Verified that..." or "Validated that..."
- Describe what the TESTER validated, not what the control performer did
- Past tense, testing perspective
- 1-2 sentences explaining HOW the validation was performed

### 3. TRANSFORMATION EXAMPLES:

**Example 1:**
Testing Attribute: "Investment Accounting Staff calculates pro-rata net assets"
✅ Test Step Name: "Inspect Calculation Process"
✅ Test Step Description: "Inspect the calculation methodology and verify the accuracy of pro-rata net asset calculations."
✅ Attribute Name: "Calculation Accuracy"
✅ Attribute Description: "Verified that pro-rata net asset calculations were performed accurately by reperforming the calculation and comparing results."

**Example 2:**
Testing Attribute: "Director reviews and approves the reconciliation"
✅ Test Step Name: "Inspect Review Evidence"
✅ Test Step Description: "Inspect evidence of management review and approval of the reconciliation."
✅ Attribute Name: "Management Review"
✅ Attribute Description: "Verified that the Director reviewed and approved the reconciliation by inspecting approval signatures and dates."

**Example 3:**
Testing Attribute: "Personnel verifies completeness and accuracy of IPE"
✅ Test Step Name: "Inspect IPE Verification"
✅ Test Step Description: "Inspect evidence that IPE was verified for completeness and accuracy."
✅ Attribute Name: "IPE Verification"
✅ Attribute Description: "Verified that IPE procedures were performed and documented by inspecting verification checklists and supporting documentation."

### 4. OUTPUT FORMAT:
Return ONLY a JSON structure:
```json
{
  "test_steps": [
    {
      "control_id": "[Use the Ref ID]",
      "name": "Test Step Name",
      "description": "Detailed description",
      "attribute_name": "Attribute Name or N/A",
      "attribute_description": "Detailed attribute description or N/A"
    }
  ]
}
```

### 5. FINAL REQUIREMENTS:
- NEVER copy testing attribute text verbatim
- ALWAYS transform process descriptions into testing instructions
- Create 5-7 test steps total (1 "Obtain Evidence" + 4-6 testing steps)
- Each test step must have a unique, concise name
- All attribute descriptions must be in past tense from tester perspective
- Focus on creating ACTIONABLE test steps that auditors can actually perform"""

        user_prompt = f"""TRANSFORM the following SOX control testing attributes into proper test step format:

Control ID: {control_data['ref_id']}
Control Description: {control_data['control_description']}
Testing Attributes: {control_data['testing_attributes']}
Design Attributes: {control_data['design_attributes']}
Evidence of Control: {control_data['evidence_of_control']}

CRITICAL INSTRUCTIONS:
1. DO NOT copy any testing attribute text verbatim
2. TRANSFORM each testing attribute (A, B, C, D, E) into a proper test step
3. Create concise test step names (2-4 words) that capture the core action
4. Rewrite descriptions as testing instructions using "Inspect", "Verify", "Review"
5. Create meaningful attribute names that describe what's being validated
6. Write attribute descriptions from the tester's perspective in past tense

Example transformation for your reference:
- Testing Attribute: "Staff obtains statements from sponsors"
- Becomes Test Step: Name="Inspect Statement Procurement", Description="Inspect evidence that statements were obtained from equity method sponsors for the selected period."

Now transform ALL the testing attributes for this control following these rules."""

    response = make_openai_request(system_prompt, user_prompt, max_tokens=2500)
    
    return {
        'control_id': control_data['ref_id'],
        'control_description': control_data['control_description'],
        'ai_generated_content': response,
        'original_testing_attributes': control_data['testing_attributes'],
        'original_design_attributes': control_data['design_attributes'],
        'original_evidence': control_data['evidence_of_control'],
        'is_na_scenario': is_na_scenario
    }

def create_excel_template(processed_controls: List[Dict]) -> str:
    """Create an Excel template with the processed test steps and attributes."""
    logger.info("Creating Excel template with processed controls")
    
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "SOX Test Steps Template"
        
        # Set up headers to match the expected output format
        headers = [
            'Control ID',
            'Test Step Name', 
            'Test Step Description',
            'Attribute Name',
            'Attribute Description'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Process each control and extract test steps
        current_row = 2
        
        for control in processed_controls:
            control_id = control['control_id']
            ai_content = control['ai_generated_content']
            
            # Try to parse the JSON response from AI
            test_steps = []
            
            try:
                # Extract JSON from the AI response
                if '```json' in ai_content:
                    json_start = ai_content.find('```json') + 7
                    json_end = ai_content.find('```', json_start)
                    json_content = ai_content[json_start:json_end].strip()
                else:
                    # Try to find JSON structure in the response
                    json_start = ai_content.find('{')
                    json_end = ai_content.rfind('}') + 1
                    json_content = ai_content[json_start:json_end]
                
                parsed_data = json.loads(json_content)
                test_steps = parsed_data.get('test_steps', [])
                
            except (json.JSONDecodeError, ValueError) as e:
                logger.warning(f"Could not parse AI JSON response for control {control_id}: {e}")
                
                # Check if this was an N/A scenario
                is_na_scenario = control.get('is_na_scenario', False)
                
                if is_na_scenario:
                    # Fallback for N/A scenarios - generic but professional test steps
                    test_steps = [
                        {
                            'control_id': control_id,
                            'name': 'Obtain Evidence',
                            'description': f"For a sample period, obtain evidence of the control described as: {control['control_description'][:100]}...",
                            'attribute_name': 'N/A',
                            'attribute_description': 'N/A'
                        },
                        {
                            'control_id': control_id,
                            'name': 'Inspect Control Design',
                            'description': 'Inspect the design of the control to understand the control objective and how it operates.',
                            'attribute_name': 'Control Design Understanding',
                            'attribute_description': 'Verified that the control design is appropriate to address the identified risk and achieve the control objective.'
                        },
                        {
                            'control_id': control_id,
                            'name': 'Inspect Control Operation',
                            'description': 'Inspect evidence that the control operated effectively during the period.',
                            'attribute_name': 'Operating Effectiveness',
                            'attribute_description': 'Verified that the control operated as designed throughout the testing period by examining supporting documentation.'
                        },
                        {
                            'control_id': control_id,
                            'name': 'Verify Control Performance',
                            'description': 'Verify that the control performer has appropriate authority and competence to execute the control.',
                            'attribute_name': 'Control Performer Competence',
                            'attribute_description': 'Verified that the control performer has the appropriate authority, training, and competence to effectively perform the control.'
                        },
                        {
                            'control_id': control_id,
                            'name': 'Inspect Documentation',
                            'description': 'Inspect the completeness and accuracy of documentation supporting the control.',
                            'attribute_name': 'Documentation Completeness',
                            'attribute_description': 'Verified that supporting documentation is complete, accurate, and provides sufficient evidence of control performance.'
                        }
                    ]
                else:
                    # Enhanced fallback for normal scenarios with full control information
                    # Try to extract some basic test steps from the testing attributes
                    original_testing_attrs = control.get('original_testing_attributes', '')
                    original_evidence = control.get('original_evidence', '')
                    
                    test_steps = [
                        {
                            'control_id': control_id,
                            'name': 'Obtain Evidence',
                            'description': f"For a sample month, obtain the following evidence: {original_evidence[:200]}{'...' if len(original_evidence) > 200 else ''}",
                            'attribute_name': 'N/A',
                            'attribute_description': 'N/A'
                        }
                    ]
                    
                    # Try to create basic test steps from testing attributes if available
                    if original_testing_attrs and len(original_testing_attrs.strip()) > 0:
                        # Split by letter markers (A), B), C), etc.)
                        import re
                        attr_pattern = r'[A-Z]\)\s*'
                        attr_items = re.split(attr_pattern, original_testing_attrs)
                        attr_items = [item.strip() for item in attr_items if item.strip()]
                        
                        for i, attr_text in enumerate(attr_items[:5]):  # Limit to 5 items
                            # Extract key action words to create test step names
                            if 'obtain' in attr_text.lower() or 'receive' in attr_text.lower():
                                step_name = 'Inspect Document Procurement'
                                description = 'Inspect evidence that required documents were obtained from appropriate sources.'
                                attr_name = 'Document Completeness'
                                attr_desc = 'Verified that all required documents were obtained and are complete.'
                            elif 'calculat' in attr_text.lower() or 'comput' in attr_text.lower():
                                step_name = 'Inspect Calculation Process'
                                description = 'Inspect calculation methodology and verify computational accuracy.'
                                attr_name = 'Calculation Accuracy'
                                attr_desc = 'Verified calculation accuracy by reperforming calculations and comparing results.'
                            elif 'reconcil' in attr_text.lower():
                                step_name = 'Inspect Reconciliation'
                                description = 'Inspect the reconciliation process and verify completeness of reconciling items.'
                                attr_name = 'Reconciliation Completeness'
                                attr_desc = 'Verified that reconciliation was complete and all variances were appropriately addressed.'
                            elif 'review' in attr_text.lower() or 'approv' in attr_text.lower():
                                step_name = 'Inspect Review Evidence'
                                description = 'Inspect evidence of management review and approval.'
                                attr_name = 'Management Review'
                                attr_desc = 'Verified that appropriate management review and approval was performed and documented.'
                            elif 'verif' in attr_text.lower() or 'validat' in attr_text.lower() or 'ipe' in attr_text.lower():
                                step_name = 'Inspect Verification Process'
                                description = 'Inspect evidence of data verification and validation procedures.'
                                attr_name = 'Data Verification'
                                attr_desc = 'Verified that data verification procedures were performed and documented appropriately.'
                            else:
                                step_name = f'Inspect Control Activity {i+1}'
                                description = f'Inspect evidence of control activity performance for the selected period.'
                                attr_name = f'Control Activity {i+1}'
                                attr_desc = 'Verified that control activity was performed as designed and documented appropriately.'
                            
                            test_steps.append({
                                'control_id': control_id,
                                'name': step_name,
                                'description': description,
                                'attribute_name': attr_name,
                                'attribute_description': attr_desc
                            })
                    else:
                        # Generic fallback test steps when no testing attributes available
                        test_steps.extend([
                            {
                                'control_id': control_id,
                                'name': 'Inspect Control Design',
                                'description': 'Inspect the design of the control to understand the control objective and how it operates.',
                                'attribute_name': 'Control Design Understanding',
                                'attribute_description': 'Verified that the control design is appropriate to address the identified risk and achieve the control objective.'
                            },
                            {
                                'control_id': control_id,
                                'name': 'Inspect Control Operation',
                                'description': 'Inspect evidence that the control operated effectively during the period.',
                                'attribute_name': 'Operating Effectiveness',
                                'attribute_description': 'Verified that the control operated as designed throughout the testing period by examining supporting documentation.'
                            },
                            {
                                'control_id': control_id,
                                'name': 'Verify Control Performance',
                                'description': 'Verify that the control performer has appropriate authority and competence to execute the control.',
                                'attribute_name': 'Control Performer Competence',
                                'attribute_description': 'Verified that the control performer has the appropriate authority, training, and competence to effectively perform the control.'
                            },
                            {
                                'control_id': control_id,
                                'name': 'Inspect Documentation',
                                'description': 'Inspect the completeness and accuracy of documentation supporting the control.',
                                'attribute_name': 'Documentation Completeness',
                                'attribute_description': 'Verified that supporting documentation is complete, accurate, and provides sufficient evidence of control performance.'
                            }
                        ])
            
            # Add each test step to the Excel
            for step in test_steps:
                # Use control_id from step if available, otherwise use the control_id from the loop
                step_control_id = step.get('control_id', control_id)
                ws.cell(row=current_row, column=1, value=step_control_id)
                ws.cell(row=current_row, column=2, value=step.get('name', ''))
                ws.cell(row=current_row, column=3, value=step.get('description', ''))
                ws.cell(row=current_row, column=4, value=step.get('attribute_name', ''))
                ws.cell(row=current_row, column=5, value=step.get('attribute_description', ''))
                current_row += 1
        
        # Save the workbook
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            output_path = temp_file.name
        
        wb.save(output_path)
        logger.info(f"Excel template created: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Error creating Excel template: {str(e)}")
        raise

def generate_test_steps(file_paths: List[str], template: str = '') -> Dict:
    """Main function to process Excel file with SOX controls and generate test steps."""
    logger.info(f"Processing SOX controls Excel file: {file_paths[0]}")
    
    try:
        # Only process the first file (should be Excel)
        excel_file_path = file_paths[0]
        
        # Parse the Excel file to extract control information
        controls = parse_sox_controls_excel(excel_file_path)
        
        if not controls:
            raise ValueError("No controls found in the Excel file")
        
        # Process each control to generate test steps
        processed_controls = []
        for control in controls:
            logger.info(f"Processing control: {control['ref_id']}")
            processed_control = generate_test_steps_from_control(control)
            processed_controls.append(processed_control)
        
        # Create Excel template with all processed controls
        excel_template_path = create_excel_template(processed_controls)
        
        # Return structured response
        result = {
            'id': datetime.now().strftime('%Y%m%d_%H%M%S'),
            'controlName': f'SOX Controls Processing - {len(controls)} controls',
            'controlsProcessed': len(controls),
            'excelTemplatePath': excel_template_path,
            'processedControls': processed_controls,
            'createdAt': datetime.now().isoformat()
        }
        
        logger.info(f"Successfully processed {len(controls)} controls")
        return result
        
    except Exception as e:
        logger.error(f"Error processing SOX controls: {str(e)}")
        raise

def export_test_plan_to_word(test_plan_data: Dict) -> str:
    """Export processed controls as Excel file (not Word for this use case)."""
    logger.info("Exporting test plan as Excel template")
    
    try:
        # If we already have an Excel template path, return it
        if 'excelTemplatePath' in test_plan_data:
            return test_plan_data['excelTemplatePath']
        
        # Otherwise, recreate the Excel template
        if 'processedControls' in test_plan_data:
            return create_excel_template(test_plan_data['processedControls'])
        
        raise ValueError("No processed controls data found")
        
    except Exception as e:
        logger.error(f"Error exporting test plan: {str(e)}")
        raise 