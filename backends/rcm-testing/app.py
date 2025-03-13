from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import tempfile
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import zipfile
import logging
import uuid
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# App configuration
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['TEMPLATE_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # limit to 16MB

# CORS configuration
CORS(app, resources={
    r"/*": {
        "origins": ["http://localhost:3000"],
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "X-Requested-With", "Accept", "Origin", "Authorization"],
        "expose_headers": ["Content-Type", "Content-Disposition"],
        "supports_credentials": False,
        "max_age": 3600
    }
})

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['TEMPLATE_FOLDER'], exist_ok=True)

def allowed_file(filename):
    """Check if uploaded file has an allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

def copy_and_customize_template(control_id, fiscal_year, testing_period, rcm_data):
    """Copy the template file and customize it with control-specific data"""
    # Path to the template file
    template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'Testing Template.xlsx')
    
    # Check if template exists
    if not os.path.exists(template_path):
        logger.error(f"Template file not found at {template_path}")
        raise FileNotFoundError(f"Template file 'Testing Template.xlsx' not found in templates directory")
    
    # Create a new workbook from the template
    wb = openpyxl.load_workbook(template_path)
    
    # Check if the Summary sheet exists in the template
    if "Summary" not in wb.sheetnames:
        logger.warning("Summary sheet not found in template. Adding a Summary sheet.")
        summary_sheet = wb.create_sheet("Summary")
    else:
        summary_sheet = wb["Summary"]
    
    # Update metadata cells in the Summary sheet
    # Assuming these cells exist in the template - adjust cell references as needed based on your template structure
    try:
        # Update the merged cell B2 with Control ID and Short Name
        short_name = rcm_data.get('short_name', '')
        summary_sheet['B2'] = f"{control_id}-{short_name}"
        
        # Direct cell assignments for specific metadata - this ensures explicit placement
        summary_sheet['C3'] = rcm_data.get('business_cycle', '')  # Business Cycle in C3
        summary_sheet['C4'] = rcm_data.get('sub_process', '')     # Sub Process in C4
        summary_sheet['C5'] = rcm_data.get('control_description', '')  # Control Description in C5
        summary_sheet['C7'] = rcm_data.get('application', '')     # Application in C7
        
        # Update RCM data cells in the table section if it exists
        # We'll track sections we've already updated to avoid duplicate entries
        updated_sections = {
            'business_cycle': True,     # Already updated in C3
            'sub_process': True,        # Already updated in C4
            'control_description': True, # Already updated in C5
            'application': True         # Already updated in C7
        }
        
        # Look for an RCM Information section table
        rcm_section_found = False
        for row in range(10, 30):  # Start searching from row 10 to avoid interfering with our fixed assignments
            for cell in summary_sheet[row]:
                if cell.value and "RCM Information" in str(cell.value):
                    # Found the RCM Information section
                    data_row = row + 3  # Assuming headers are 1 row below and data 2 rows below that
                    
                    # Try to find the headers in the row below
                    header_row = row + 2
                    header_cols = {}
                    
                    for col in range(1, 10):
                        header_val = summary_sheet.cell(row=header_row, column=col).value
                        if not header_val:
                            continue
                            
                        header_val = str(header_val).lower()
                        if "business cycle" in header_val:
                            header_cols['business_cycle'] = col
                        elif "sub process" in header_val:
                            header_cols['sub_process'] = col
                        elif "control description" in header_val:
                            header_cols['control_description'] = col
                        elif "application" in header_val:
                            header_cols['application'] = col
                    
                    # Update the cells in the data row
                    if 'business_cycle' in header_cols:
                        summary_sheet.cell(row=data_row, column=header_cols['business_cycle']).value = rcm_data.get('business_cycle', '')
                    
                    if 'sub_process' in header_cols:
                        summary_sheet.cell(row=data_row, column=header_cols['sub_process']).value = rcm_data.get('sub_process', '')
                    
                    if 'control_description' in header_cols:
                        summary_sheet.cell(row=data_row, column=header_cols['control_description']).value = rcm_data.get('control_description', '')
                    
                    if 'application' in header_cols:
                        summary_sheet.cell(row=data_row, column=header_cols['application']).value = rcm_data.get('application', '')
                    
                    rcm_section_found = True
                    break
            if rcm_section_found:
                break

    except Exception as e:
        logger.error(f"Error customizing template: {str(e)}")
        # Continue despite errors in template customization
    
    return wb

@app.route('/api/generate-testing', methods=['POST'])
def generate_testing():
    """Process the uploaded RCM template and generate testing templates"""
    # Check if a file was uploaded
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
    
    if not allowed_file(file.filename):
        return jsonify({"error": "File type not allowed. Please upload an Excel file (.xlsx, .xls)"}), 400
    
    try:
        # Check if template file exists
        template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'Testing Template.xlsx')
        if not os.path.exists(template_path):
            return jsonify({"error": "Testing template file not found. Please add 'Testing Template.xlsx' to the templates directory."}), 500
            
        # Create a unique filename for the upload
        unique_filename = str(uuid.uuid4()) + '.xlsx'
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(file_path)
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Create a temporary directory for generating files
        temp_dir = tempfile.mkdtemp()
        testing_dir = os.path.join(temp_dir, "Testing")
        os.makedirs(testing_dir, exist_ok=True)
        
        # Get fiscal year and testing period from columns B & C
        fiscal_year = None
        testing_period = None
        
        # Dictionary to track unique control IDs
        control_ids = {}
        
        # Process the RCM template
        for row in range(2, ws.max_row + 1):  # Assuming row 1 is header
            # Check if the cell in Column A is not empty
            control_id_cell = ws.cell(row=row, column=1)
            if control_id_cell.value:
                control_id = str(control_id_cell.value).strip()
                
                # Get fiscal year from column B if not already set
                if not fiscal_year:
                    fiscal_year_cell = ws.cell(row=row, column=2)
                    if fiscal_year_cell.value:
                        fiscal_year = str(fiscal_year_cell.value).strip()
                
                # Get testing period from column C if not already set
                if not testing_period:
                    testing_period_cell = ws.cell(row=row, column=3)
                    if testing_period_cell.value:
                        testing_period = str(testing_period_cell.value).strip()
                
                # Skip if we've already processed this control ID
                if control_id in control_ids:
                    continue
                
                # Extract data from the RCM template for this control
                rcm_data = {
                    'business_cycle': ws.cell(row=row, column=4).value,  # Column D
                    'sub_process': ws.cell(row=row, column=5).value,     # Column E
                    'control_description': ws.cell(row=row, column=8).value,  # Column H
                    'application': ws.cell(row=row, column=6).value,      # Column F
                    'short_name': ws.cell(row=row, column=7).value       # Column G
                }
                
                # Create folder for this control ID
                control_dir = os.path.join(testing_dir, control_id)
                os.makedirs(control_dir, exist_ok=True)
                
                # Create a Supporting Evidence folder inside the control folder
                evidence_dir = os.path.join(control_dir, "Supporting Evidence")
                os.makedirs(evidence_dir, exist_ok=True)
                
                # Add a placeholder file in the Supporting Evidence folder to ensure it's included in the zip
                placeholder_path = os.path.join(evidence_dir, "README.txt")
                with open(placeholder_path, 'w') as placeholder_file:
                    placeholder_file.write("This folder is for storing supporting evidence files for control testing.")
                
                # Create testing template for this control ID
                if fiscal_year and testing_period:
                    template_filename = f"{fiscal_year}-{control_id}-{testing_period}.xlsx"
                else:
                    template_filename = f"{control_id}-testing.xlsx"
                
                template_path = os.path.join(control_dir, template_filename)
                
                # Copy and customize the template
                try:
                    testing_wb = copy_and_customize_template(control_id, fiscal_year, testing_period, rcm_data)
                    testing_wb.save(template_path)
                    
                    # Add to processed controls
                    control_ids[control_id] = True
                except Exception as e:
                    logger.error(f"Error processing control ID {control_id}: {str(e)}")
                    # Continue with other controls despite error
        
        # Create zip file of the testing directory
        zip_path = os.path.join(temp_dir, "testing_templates.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(testing_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zipf.write(file_path, arcname)
        
        # Send the zip file
        response = send_file(
            zip_path, 
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"testing_templates_{datetime.now().strftime('%Y%m%d%H%M%S')}.zip"
        )
        
        # Clean up
        os.remove(file_path)  # Remove original uploaded file
        
        # Add cleanup callback to delete temp directory after sending file
        @response.call_on_close
        def cleanup():
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                logger.error(f"Error cleaning up temp directory: {str(e)}")
        
        return response
        
    except Exception as e:
        logger.error(f"Error generating testing templates: {str(e)}")
        logger.error(f"Traceback: {logging.traceback.format_exc()}")
        return jsonify({"error": f"Error generating testing templates: {str(e)}"}), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({"status": "healthy"}), 200

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3003, debug=True) 